"""
Automatically fill the ground truth template with predicted values
This helps speed up manual labeling by providing initial estimates
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# EU countries list
EU_COUNTRIES = [
    "Germany", "France", "Italy", "Spain", "Netherlands", "Belgium", "Austria", 
    "Sweden", "Denmark", "Finland", "Poland", "Czech Republic", "Hungary", 
    "Romania", "Bulgaria", "Greece", "Portugal", "Ireland", "Croatia", 
    "Slovenia", "Slovakia", "Lithuania", "Latvia", "Estonia", "Cyprus", 
    "Malta", "Luxembourg"
]

# CEFR level hierarchy for comparison
CEFR_LEVELS = {"A1": 1, "A2": 2, "B1": 3, "B2": 4, "C1": 5, "C2": 6}

def compare_cefr_levels(student_level, required_level):
    """Compare CEFR levels. Returns True if student meets requirement."""
    if required_level == "None" or required_level == "Not required" or not required_level:
        return True
    if student_level == "N/A" or student_level == "None" or not student_level:
        return False
    
    # Convert to string to handle any type
    student_level = str(student_level)
    required_level = str(required_level)
    
    # Extract level from strings like "IELTS: 7.0" or "Level: C1"
    if ":" in student_level:
        parts = student_level.split(":")
        if "Level" in parts[0]:
            level_part = parts[1].strip()
            if level_part == "N/A":
                return False
            student_level = level_part
        else:
            # It's a test score, need to map to CEFR
            test_name = parts[0].strip()
            score_str = parts[1].strip()
            
            # Handle N/A scores
            if score_str == "N/A" or score_str == "None":
                return False
            
            try:
                score = float(score_str)
            except ValueError:
                return False
            
            # Simplified mapping: IELTS 6.5+ = B2, 7.0+ = C1
            if "IELTS" in test_name:
                if score >= 7.5: student_level = "C1"
                elif score >= 6.5: student_level = "B2"
                elif score >= 5.5: student_level = "B1"
                else: student_level = "A2"
            elif "TOEFL" in test_name:
                if score >= 95: student_level = "C1"
                elif score >= 72: student_level = "B2"
                else: student_level = "B1"
            elif "Cambridge CAE" in test_name or "CAE" in test_name:
                student_level = "C1"
            elif "TOEIC" in test_name:
                if score >= 850: student_level = "C1"
                elif score >= 700: student_level = "B2"
                else: student_level = "B1"
            else:
                return True  # Assume pass if we can't determine
    
    student_val = CEFR_LEVELS.get(student_level, 0)
    required_val = CEFR_LEVELS.get(required_level, 0)
    
    return student_val >= required_val

def check_degree_match(student_field, required_domains):
    """Estimate degree compatibility score."""
    if not required_domains or required_domains == "":
        return "HIGH", "0.8-1.0"
    
    student_lower = student_field.lower()
    domains_lower = required_domains.lower()
    
    # Direct matches
    if student_lower in domains_lower or any(domain.strip() in student_lower for domain in domains_lower.split(",")):
        return "HIGH", "0.8-1.0"
    
    # Related field mappings
    related_fields = {
        "computer science": ["data science", "artificial intelligence", "information science", "software", "informatics"],
        "business administration": ["management", "economics", "business", "marketing", "mba"],
        "economics": ["business administration", "finance", "management"],
        "marketing": ["business", "management", "digital marketing"],
        "business informatics": ["computer science", "information systems", "digital business"]
    }
    
    for key, related in related_fields.items():
        if key in student_lower:
            if any(r in domains_lower for r in related):
                return "MEDIUM", "0.5-0.7"
    
    # Check if in same broad category
    stem_fields = ["computer", "engineering", "mathematics", "physics", "chemistry", "biology"]
    business_fields = ["business", "economics", "management", "finance", "marketing"]
    
    student_is_stem = any(f in student_lower for f in stem_fields)
    domain_is_stem = any(f in domains_lower for f in stem_fields)
    student_is_business = any(f in student_lower for f in business_fields)
    domain_is_business = any(f in domains_lower for f in business_fields)
    
    if (student_is_stem and domain_is_stem) or (student_is_business and domain_is_business):
        return "LOW", "0.3-0.5"
    
    return "NO", "0.0-0.2"

def calculate_relevance(student_interests, student_desired, program_name, course_summary):
    """Calculate relevance score for ranking."""
    score = 1  # Start with minimum
    
    program_lower = program_name.lower()
    summary_lower = course_summary.lower()
    
    # Check desired programs
    if student_desired:
        desired_list = [d.strip().lower() for d in student_desired.split(",")]
        for desired in desired_list:
            if desired in program_lower or desired in summary_lower:
                score = max(score, 5)  # Perfect match
                break
    
    # Check interests
    if student_interests:
        interest_list = [i.strip().lower() for i in student_interests.split(",")]
        matches = sum(1 for interest in interest_list if interest in program_lower or interest in summary_lower)
        
        if matches >= 2:
            score = max(score, 4)  # High relevance
        elif matches == 1:
            score = max(score, 3)  # Medium relevance
    
    return score

def estimate_top_n(relevance_score, passes_hard_constraints, degree_match):
    """Estimate where program should appear in rankings."""
    if not passes_hard_constraints:
        return "Not in Top 20"
    
    if degree_match in ["NO", "LOW"]:
        return "Not in Top 20"
    
    if relevance_score == 5 and degree_match == "HIGH":
        return "Top 3"
    elif relevance_score >= 4 and degree_match in ["HIGH", "MEDIUM"]:
        return "Top 5"
    elif relevance_score >= 3:
        return "Top 10"
    else:
        return "Top 20"

# Load data
print("Loading data...")
with open('test_profiles.json', 'r') as f:
    profiles_data = json.load(f)
    profiles = profiles_data['test_profiles']

with open('test_sample_programs.json', 'r') as f:
    programs_data = json.load(f)
    programs = programs_data['programs']

# Load Excel
print("Loading Excel template...")
wb = load_workbook('agent3_ground_truth_template.xlsx')

# Create a mapping for quick lookup
profile_map = {p['id']: p for p in profiles}
program_map = {p['program_id']: p for p in programs}

# ==========================================
# SHEET 1: Hard Constraints
# ==========================================
print("\nFilling Sheet 1: Hard Constraints...")
ws_hard = wb["Hard Constraints"]

for row_idx in range(2, ws_hard.max_row + 1):
    profile_id = ws_hard.cell(row_idx, 1).value
    program_id = ws_hard.cell(row_idx, 3).value
    
    if not profile_id or not program_id:
        continue
    
    profile = profile_map.get(profile_id)
    program = program_map.get(program_id)
    
    if not profile or not program:
        continue
    
    gold = profile['gold_standard']['expected_profile']
    
    # GPA Check
    student_gpa = ws_hard.cell(row_idx, 6).value
    program_gpa = ws_hard.cell(row_idx, 7).value
    if program_gpa and student_gpa:
        gpa_pass = "YES" if student_gpa <= (program_gpa + 0.05) else "NO"
    else:
        gpa_pass = "YES"
    ws_hard.cell(row_idx, 8).value = gpa_pass
    
    # Tuition Check
    student_budget = ws_hard.cell(row_idx, 9).value
    tuition_eu = ws_hard.cell(row_idx, 10).value or 0
    tuition_non_eu = ws_hard.cell(row_idx, 11).value or 0
    
    citizenship = gold.get('citizenship', '')
    is_eu = citizenship in EU_COUNTRIES
    applicable_tuition = tuition_eu if is_eu else tuition_non_eu
    
    if student_budget and student_budget > 0:
        tuition_pass = "YES" if applicable_tuition <= (student_budget + 100) else "NO"
    else:
        tuition_pass = "YES"
    ws_hard.cell(row_idx, 12).value = tuition_pass
    
    # Location Check (Cities)
    pref_cities_str = ws_hard.cell(row_idx, 13).value
    program_city = ws_hard.cell(row_idx, 14).value
    
    if pref_cities_str and pref_cities_str != "None" and program_city:
        pref_cities = [c.strip().lower() for c in pref_cities_str.split(",")]
        location_pass = "YES" if program_city.lower() in pref_cities else "NO"
    else:
        location_pass = "YES"
    ws_hard.cell(row_idx, 15).value = location_pass
    
    # English Check
    student_english = ws_hard.cell(row_idx, 16).value
    program_english = ws_hard.cell(row_idx, 17).value
    
    english_pass = "YES" if compare_cefr_levels(student_english, program_english) else "NO"
    ws_hard.cell(row_idx, 18).value = english_pass
    
    # Work Experience Check
    student_exp = ws_hard.cell(row_idx, 19).value or 0
    program_exp = ws_hard.cell(row_idx, 20).value or 0
    work_pass = "YES" if student_exp >= program_exp else "NO"
    ws_hard.cell(row_idx, 21).value = work_pass
    
    # Semester Check
    pref_semester = ws_hard.cell(row_idx, 22).value
    winter_avail = ws_hard.cell(row_idx, 23).value
    summer_avail = ws_hard.cell(row_idx, 24).value
    
    if pref_semester:
        if "winter" in pref_semester.lower():
            semester_pass = "YES" if winter_avail == "Yes" else "NO"
        elif "summer" in pref_semester.lower():
            semester_pass = "YES" if summer_avail == "Yes" else "NO"
        else:
            semester_pass = "YES"
    else:
        semester_pass = "YES"
    ws_hard.cell(row_idx, 25).value = semester_pass
    
    # ECTS Check
    student_ects = ws_hard.cell(row_idx, 26).value or 0
    program_ects = ws_hard.cell(row_idx, 27).value or 180
    ects_pass = "YES" if student_ects >= program_ects else "NO"
    ws_hard.cell(row_idx, 28).value = ects_pass
    
    # Overall Pass
    all_checks = [gpa_pass, tuition_pass, location_pass, english_pass, work_pass, semester_pass, ects_pass]
    overall_pass = "YES" if all(check == "YES" for check in all_checks) else "NO"
    ws_hard.cell(row_idx, 29).value = overall_pass
    
    # Add note if failed
    if overall_pass == "NO":
        failed = [name for name, val in zip(
            ["GPA", "Tuition", "Location", "English", "Work Exp", "Semester", "ECTS"],
            all_checks
        ) if val == "NO"]
        ws_hard.cell(row_idx, 30).value = f"Failed: {', '.join(failed)}"

print(f"  Filled {ws_hard.max_row - 1} rows")

# ==========================================
# SHEET 2: Degree Compatibility
# ==========================================
print("\nFilling Sheet 2: Degree Compatibility...")
ws_degree = wb["Degree Compatibility"]

for row_idx in range(2, ws_degree.max_row + 1):
    profile_id = ws_degree.cell(row_idx, 1).value
    program_id = ws_degree.cell(row_idx, 2).value
    
    if not profile_id or not program_id:
        continue
    
    student_field = ws_degree.cell(row_idx, 4).value
    required_domains = ws_degree.cell(row_idx, 5).value
    
    match_level, score_range = check_degree_match(student_field, required_domains)
    
    ws_degree.cell(row_idx, 6).value = match_level
    ws_degree.cell(row_idx, 7).value = score_range
    
    # Add reasoning
    if match_level == "HIGH":
        ws_degree.cell(row_idx, 8).value = "Direct or very close match"
    elif match_level == "MEDIUM":
        ws_degree.cell(row_idx, 8).value = "Related field with overlap"
    elif match_level == "LOW":
        ws_degree.cell(row_idx, 8).value = "Same broad category but different focus"
    else:
        ws_degree.cell(row_idx, 8).value = "Unrelated fields"

print(f"  Filled {ws_degree.max_row - 1} rows")

# ==========================================
# SHEET 3: Overall Ranking
# ==========================================
print("\nFilling Sheet 3: Overall Ranking...")
ws_ranking = wb["Overall Ranking"]

# First pass: calculate all relevance scores per profile
profile_rankings = {}

for row_idx in range(2, ws_ranking.max_row + 1):
    profile_id = ws_ranking.cell(row_idx, 1).value
    program_id = ws_ranking.cell(row_idx, 2).value
    
    if not profile_id or not program_id:
        continue
    
    program_name = ws_ranking.cell(row_idx, 3).value
    course_summary = ws_ranking.cell(row_idx, 4).value or ""
    student_interests = ws_ranking.cell(row_idx, 5).value
    student_desired = ws_ranking.cell(row_idx, 6).value
    
    # Get hard constraint result
    profile = profile_map.get(profile_id)
    program = program_map.get(program_id)
    gold = profile['gold_standard']['expected_profile']
    
    # Find corresponding row in hard constraints sheet
    passes_hard = False
    for hc_row in range(2, ws_hard.max_row + 1):
        if (ws_hard.cell(hc_row, 1).value == profile_id and 
            ws_hard.cell(hc_row, 3).value == program_id):
            passes_hard = ws_hard.cell(hc_row, 29).value == "YES"
            break
    
    # Find degree match
    degree_match = "MEDIUM"
    for deg_row in range(2, ws_degree.max_row + 1):
        if (ws_degree.cell(deg_row, 1).value == profile_id and 
            ws_degree.cell(deg_row, 2).value == program_id):
            degree_match = ws_degree.cell(deg_row, 6).value
            break
    
    relevance = calculate_relevance(student_interests, student_desired, program_name, course_summary)
    
    ws_ranking.cell(row_idx, 7).value = relevance
    
    # Store for ranking calculation
    if profile_id not in profile_rankings:
        profile_rankings[profile_id] = []
    
    profile_rankings[profile_id].append({
        'row': row_idx,
        'relevance': relevance,
        'passes_hard': passes_hard,
        'degree_match': degree_match,
        'program_id': program_id
    })

# Second pass: assign Top N based on ranking within each profile
for profile_id, programs_list in profile_rankings.items():
    # Sort by relevance (descending), then by passes_hard
    programs_list.sort(key=lambda x: (x['passes_hard'], x['relevance']), reverse=True)
    
    for rank, prog_info in enumerate(programs_list, 1):
        row_idx = prog_info['row']
        top_n = estimate_top_n(prog_info['relevance'], prog_info['passes_hard'], prog_info['degree_match'])
        
        # Refine based on actual rank
        if prog_info['passes_hard']:
            if rank <= 3:
                top_n = "Top 3"
            elif rank <= 5:
                top_n = "Top 5"
            elif rank <= 10:
                top_n = "Top 10"
            else:
                top_n = "Top 20"
        else:
            top_n = "Not in Top 20"
        
        ws_ranking.cell(row_idx, 8).value = top_n
        
        # Add notes
        if not prog_info['passes_hard']:
            ws_ranking.cell(row_idx, 9).value = "Failed hard constraints"
        elif prog_info['degree_match'] == "NO":
            ws_ranking.cell(row_idx, 9).value = "Degree mismatch"

print(f"  Filled {ws_ranking.max_row - 1} rows")

# Save
output_file = 'agent3_ground_truth_FILLED.xlsx'
wb.save(output_file)

print(f"\n✅ Auto-filled template saved as: {output_file}")
print(f"\nSummary:")
print(f"  - All 100 combinations have been filled")
print(f"  - Sheet 1: Hard constraints evaluated")
print(f"  - Sheet 2: Degree compatibility assessed")
print(f"  - Sheet 3: Relevance scores and rankings assigned")
print(f"\nNext steps:")
print(f"  1. Open {output_file}")
print(f"  2. Review and adjust the auto-filled values")
print(f"  3. Save as 'agent3_ground_truth.xlsx' when satisfied")
print(f"  4. Run 'test_agent3_with_ground_truth.py'")
