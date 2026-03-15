"""
Filter agent3_ground_truth_FILLED.xlsx to remove:
- Sheet 2 (Degree Compatibility): Keep only programs where 'Overall Should Pass?' is:
  Empty, "All constraints passed", OR contains only "ECTS Domain:..." (only ECTS needs manual check)
- Sheet 3 (Overall Ranking): Programs with relevance score <= 0.5 (on 1-5 scale)
"""

import openpyxl
from collections import defaultdict

def filter_ground_truth_sheets(input_file, output_file):
    """Filter sheets based on hard constraints and relevance scores"""
    
    wb = openpyxl.load_workbook(input_file)
    
    # Step 1: Identify programs that PASS all hard constraints
    ws_hard = wb['Hard Constraints']
    
    # Get column indices for Hard Constraints sheet
    headers_hard = [cell.value for cell in ws_hard[1]]
    col_profile = headers_hard.index('Profile ID') + 1
    col_program_id = headers_hard.index('Program ID') + 1
    col_overall = headers_hard.index('Overall Should Pass?') + 1
    
    # Track which (profile, program) combinations should be included
    # Include if 'Overall Should Pass?' is empty OR contains only "ECTS Domain:"
    programs_pass_hard_constraints = set()
    
    print("Analyzing hard constraints using 'Overall Should Pass?' column...")
    for row in ws_hard.iter_rows(min_row=2, max_row=ws_hard.max_row):
        profile_id = row[col_profile - 1].value
        program_id = row[col_program_id - 1].value
        overall_status = row[col_overall - 1].value
        
        if not profile_id or not program_id:
            continue
        
        # Include if:
        # 1. 'Overall Should Pass?' is empty/None (all constraints passed) OR
        # 2. 'Overall Should Pass?' says "All constraints passed" OR
        # 3. 'Overall Should Pass?' contains ONLY "ECTS Domain:..." (only ECTS needs manual check)
        #    (not combined with other failures like "Location: ...")
        should_include = False
        
        if not overall_status:
            # Empty means all constraints passed
            should_include = True
        else:
            overall_str = str(overall_status).strip()
            # Include if explicitly says "All constraints passed"
            if overall_str == "All constraints passed":
                should_include = True
            # OR if it ONLY mentions ECTS Domain (no other failures indicated by "|")
            elif "ECTS Domain:" in overall_str and "|" not in overall_str:
                should_include = True
        
        if should_include:
            programs_pass_hard_constraints.add((profile_id, str(program_id)))
    
    print(f"✓ Found {len(programs_pass_hard_constraints)} (profile, program) combinations that pass ALL hard constraints")
    
    # Step 2: Filter Sheet 2 (Degree Compatibility) - keep only programs that pass hard constraints
    ws_degree = wb['Degree Compatibility']
    headers_deg = [cell.value for cell in ws_degree[1]]
    col_profile_deg = headers_deg.index('Profile ID') + 1
    col_program_deg = headers_deg.index('Program ID') + 1
    
    rows_to_delete_deg = []
    kept_deg = 0
    deleted_deg = 0
    
    print("\nFiltering Sheet 2 (Degree Compatibility)...")
    for row_idx in range(2, ws_degree.max_row + 1):
        row = ws_degree[row_idx]
        profile_id = row[col_profile_deg - 1].value
        program_id = row[col_program_deg - 1].value
        
        if not profile_id or not program_id:
            continue
        
        # Keep only if this program passes ALL hard constraints
        if (profile_id, str(program_id)) in programs_pass_hard_constraints:
            kept_deg += 1
        else:
            rows_to_delete_deg.append(row_idx)
            deleted_deg += 1
    
    # Delete rows in reverse order to avoid index shifting
    for row_idx in reversed(rows_to_delete_deg):
        ws_degree.delete_rows(row_idx, 1)
    
    print(f"  Kept: {kept_deg} rows")
    print(f"  Deleted: {deleted_deg} rows")
    print(f"  New total: {ws_degree.max_row - 1} rows (excluding header)")
    
    # Step 3: Filter Sheet 3 (Overall Ranking) - keep only programs with relevance score >= 2.5 (0.5 on normalized scale)
    ws_ranking = wb['Overall Ranking']
    headers_rank = [cell.value for cell in ws_ranking[1]]
    col_profile_rank = headers_rank.index('Profile ID') + 1
    col_program_rank = headers_rank.index('Program ID') + 1
    col_relevance = headers_rank.index('Relevance Score (1-5)') + 1
    
    rows_to_delete_rank = []
    kept_rank = 0
    deleted_rank = 0
    
    print("\nFiltering Sheet 3 (Overall Ranking)...")
    for row_idx in range(2, ws_ranking.max_row + 1):
        row = ws_ranking[row_idx]
        profile_id = row[col_profile_rank - 1].value
        program_id = row[col_program_rank - 1].value
        relevance_score = row[col_relevance - 1].value
        
        if not profile_id or not program_id:
            continue
        
        # Keep only if:
        # 1. Program passes ALL hard constraints (except ECTS) AND
        # 2. Relevance score > 0.5 (on 1-5 scale)
        passes_hard = (profile_id, str(program_id)) in programs_pass_hard_constraints
        
        # Parse relevance score
        try:
            if relevance_score is not None:
                score = float(relevance_score)
            else:
                score = 0.0
        except (ValueError, TypeError):
            score = 0.0
        
        if passes_hard and score > 0.5:
            kept_rank += 1
        else:
            rows_to_delete_rank.append(row_idx)
            deleted_rank += 1
    
    # Delete rows in reverse order
    for row_idx in reversed(rows_to_delete_rank):
        ws_ranking.delete_rows(row_idx, 1)
    
    print(f"  Kept: {kept_rank} rows")
    print(f"  Deleted: {deleted_rank} rows")
    print(f"  New total: {ws_ranking.max_row - 1} rows (excluding header)")
    
    # Save the filtered workbook
    wb.save(output_file)
    print(f"\n✓ Filtered workbook saved to: {output_file}")
    
    # Summary
    print("\n" + "="*80)
    print("FILTERING SUMMARY")
    print("="*80)
    print(f"Programs passing ALL hard constraints: {len(programs_pass_hard_constraints)}")
    print(f"\nSheet 2 (Degree Compatibility):")
    print(f"  Before: {kept_deg + deleted_deg} rows")
    print(f"  After:  {kept_deg} rows")
    print(f"  Removed: {deleted_deg} rows")
    print(f"\nSheet 3 (Overall Ranking):")
    print(f"  Before: {kept_rank + deleted_rank} rows")
    print(f"  After:  {kept_rank} rows")
    print(f"  Removed: {deleted_rank} rows")
    print(f"  Filter criteria: Notes contains 'All constraints passed' or 'ECTS Domain:...' AND relevance score > 0.5/5.0")

if __name__ == '__main__':
    input_file = 'agent3_ground_truth_FILLED.xlsx'
    output_file = 'agent3_ground_truth_FILLED.xlsx'  # Overwrite the same file
    
    print("="*80)
    print("FILTERING GROUND TRUTH EXCEL FILE")
    print("="*80)
    print(f"Input:  {input_file}")
    print(f"Output: {output_file}")
    print("\nFiltering criteria:")
    print("  Sheet 2 (Degree Compatibility): Keep programs where 'Overall Should Pass?' is:")
    print("                                   Empty, 'All constraints passed', OR 'ECTS Domain:...' only")
    print("  Sheet 3 (Overall Ranking): Keep only programs with relevance score > 0.5/5.0")
    print("                             AND matching Sheet 2 criteria")
    print("="*80 + "\n")
    
    filter_ground_truth_sheets(input_file, output_file)
