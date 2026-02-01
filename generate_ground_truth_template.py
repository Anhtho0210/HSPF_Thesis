"""
Generate Excel template for Agent3 ground truth labeling
Creates a structured template with all 5 profiles × 20 programs combinations
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Load test data
with open('test_profiles.json', 'r') as f:
    profiles_data = json.load(f)
    profiles = profiles_data['test_profiles']

with open('test_sample_programs.json', 'r') as f:
    programs_data = json.load(f)
    programs = programs_data['programs']

print(f"Loaded {len(profiles)} profiles and {len(programs)} programs")

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ==========================================
# SHEET 1: Hard Constraints
# ==========================================
print("\nCreating Sheet 1: Hard Constraints...")

hard_constraints_data = []

for profile in profiles:
    profile_id = profile['id']
    profile_desc = profile['description']
    gold = profile['gold_standard']['expected_profile']
    
    for program in programs:
        program_id = program['program_id']
        program_name = program['program_name']
        university = program['university_name']
        
        # Pre-fill some obvious data for guidance
        tuition_eu = program.get('tuition_fee_per_semester_eur', 0)
        tuition_non_eu = program.get('non_eu_tuition_fee_eur') or 0
        min_gpa = program.get('min_gpa_german_scale')
        city = program.get('city', '')
        
        # Get student's preferred cities
        preferred_cities = gold.get('preferred_cities', [])
        preferred_cities_str = ', '.join(preferred_cities) if preferred_cities else 'None'
        
        # Get student's English proficiency
        english_test = gold.get('english_test', 'None')
        english_score = gold.get('english_score', 'N/A')
        english_level = gold.get('english_level', 'N/A')
        student_english = f"{english_test}: {english_score}" if english_test != 'None' else f"Level: {english_level}"
        
        # Get program's English requirement
        english_req = program.get('english_requirements', {})
        prog_english_level = english_req.get('min_cefr_level', 'Not specified')
        if prog_english_level == 'None':
            prog_english_level = 'Not required'
        
        # Determine if should pass (leave blank for manual labeling)
        hard_constraints_data.append({
            'Profile ID': profile_id,
            'Profile Description': profile_desc,
            'Program ID': program_id,
            'Program Name': program_name,
            'University': university,
            'Student GPA (German)': gold.get('gpa_german'),
            'Program Min GPA': min_gpa,
            'GPA Pass?': '',  # To be filled manually
            'Student Max Tuition': gold.get('max_tuition'),
            'Program Tuition (EU)': tuition_eu,
            'Program Tuition (Non-EU)': tuition_non_eu,
            'Tuition Pass?': '',  # To be filled manually
            'Student Preferred Cities': preferred_cities_str,
            'Program City': city,
            'Location Pass?': '',  # To be filled manually
            'Student English': student_english,
            'Program English Requirement': prog_english_level,
            'English Pass?': '',  # To be filled manually
            'Student Work Exp (months)': gold.get('work_experience_months', 0),
            'Program Min Work Exp': program.get('min_work_experience_months', 0),
            'Work Exp Pass?': '',  # To be filled manually
            'Student Preferred Semester': gold.get('preferred_semester'),
            'Winter Available?': 'Yes' if program.get('deadlines', {}).get('winter_semester') else 'No',
            'Summer Available?': 'Yes' if program.get('deadlines', {}).get('summer_semester') else 'No',
            'Semester Pass?': '',  # To be filled manually
            'Student ECTS': gold.get('expected_ects'),
            'Program Min ECTS': program.get('min_degree_ects', 180),
            'ECTS Pass?': '',  # To be filled manually
            'Overall Should Pass?': '',  # To be filled manually
            'Notes': ''
        })

df_hard = pd.DataFrame(hard_constraints_data)
ws_hard = wb.create_sheet("Hard Constraints")

# Write headers with formatting
for r_idx, row in enumerate(dataframe_to_rows(df_hard, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_hard.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:  # Header row
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Adjust column widths
for column in ws_hard.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 40)
    ws_hard.column_dimensions[column_letter].width = adjusted_width

print(f"  Created {len(hard_constraints_data)} rows (5 profiles × 20 programs)")

# ==========================================
# SHEET 2: Degree Compatibility
# ==========================================
print("\nCreating Sheet 2: Degree Compatibility...")

degree_data = []

for profile in profiles:
    profile_id = profile['id']
    gold = profile['gold_standard']['expected_profile']
    bachelor_field = gold.get('bachelor_field', '')
    
    for program in programs:
        program_id = program['program_id']
        program_name = program['program_name']
        required_domains = ', '.join(program.get('required_degree_domains', [])[:3])  # First 3
        
        degree_data.append({
            'Profile ID': profile_id,
            'Program ID': program_id,
            'Program Name': program_name,
            'Student Bachelor Field': bachelor_field,
            'Required Degree Domains': required_domains,
            'Should Match?': '',  # HIGH/MEDIUM/LOW/NO
            'Expected Score Range': '',  # e.g., "0.8-1.0"
            'Notes': ''
        })

df_degree = pd.DataFrame(degree_data)
ws_degree = wb.create_sheet("Degree Compatibility")

for r_idx, row in enumerate(dataframe_to_rows(df_degree, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_degree.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for column in ws_degree.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws_degree.column_dimensions[column_letter].width = adjusted_width

print(f"  Created {len(degree_data)} rows")

# ==========================================
# SHEET 3: Overall Ranking
# ==========================================
print("\nCreating Sheet 3: Overall Ranking...")

ranking_data = []

for profile in profiles:
    profile_id = profile['id']
    gold = profile['gold_standard']['expected_profile']
    interests = ', '.join(gold.get('interests', [])[:3])
    desired_programs = ', '.join(gold.get('desired_programs', []))
    
    for program in programs:
        program_id = program['program_id']
        program_name = program['program_name']
        course_summary = program.get('course_content_summary', '')
        
        ranking_data.append({
            'Profile ID': profile_id,
            'Program ID': program_id,
            'Program Name': program_name,
            'Course Content Summary': course_summary,
            'Student Interests': interests,
            'Student Desired Programs': desired_programs,
            'Relevance Score (1-5)': '',  # 5=Perfect, 4=High, 3=Medium, 2=Low, 1=None
            'Expected Top N': '',  # e.g., "1-3", "Top 5", "Top 10", "Not in Top 20"
            'Notes': ''
        })

df_ranking = pd.DataFrame(ranking_data)
ws_ranking = wb.create_sheet("Overall Ranking")

for r_idx, row in enumerate(dataframe_to_rows(df_ranking, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_ranking.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for column in ws_ranking.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws_ranking.column_dimensions[column_letter].width = adjusted_width

print(f"  Created {len(ranking_data)} rows")

# ==========================================
# SHEET 4: Instructions
# ==========================================
print("\nCreating Sheet 4: Instructions...")

ws_instructions = wb.create_sheet("Instructions", 0)  # Make it first sheet

instructions = [
    ["Agent 3 Ground Truth Labeling Template"],
    [""],
    ["Purpose:", "Manually label expected outcomes for Agent3 filtering and ranking"],
    [""],
    ["Instructions:"],
    [""],
    ["Sheet 1: Hard Constraints"],
    ["  1. Review student requirements (GPA, tuition budget, preferred cities, English proficiency, work exp, semester, ECTS)"],
    ["  2. Review program requirements (including both EU and Non-EU tuition fees)"],
    ["  3. For each constraint, mark 'YES' or 'NO' in the '...Pass?' columns"],
    ["  4. Note: Location check now uses preferred cities instead of state"],
    ["  5. Note: English language proficiency is now explicitly checked"],
    ["  6. Mark 'Overall Should Pass?' as 'YES' only if ALL constraints pass"],
    ["  7. Add notes for edge cases or unclear situations"],
    [""],
    ["Sheet 2: Degree Compatibility"],
    ["  1. Compare student's bachelor field with program's required degree domains"],
    ["  2. Mark 'Should Match?' as:"],
    ["     - HIGH: Direct match (e.g., CS student → CS program)"],
    ["     - MEDIUM: Related field (e.g., CS student → Data Science)"],
    ["     - LOW: Somewhat related (e.g., Business student → Economics program)"],
    ["     - NO: Unrelated (e.g., CS student → Agriculture)"],
    ["  3. Estimate 'Expected Score Range' (0.0-1.0):"],
    ["     - HIGH: 0.8-1.0"],
    ["     - MEDIUM: 0.5-0.7"],
    ["     - LOW: 0.3-0.5"],
    ["     - NO: 0.0-0.2"],
    [""],
    ["Sheet 3: Overall Ranking"],
    ["  1. Consider student's interests and desired programs"],
    ["  2. Assign 'Relevance Score (1-5)':"],
    ["     - 5: Perfect match (interests + field + requirements)"],
    ["     - 4: High relevance (most criteria match)"],
    ["     - 3: Medium relevance (some criteria match)"],
    ["     - 2: Low relevance (minimal match)"],
    ["     - 1: No relevance (completely unrelated)"],
    ["  3. Specify 'Expected Top N' (e.g., '1-3', 'Top 5', 'Top 10', 'Not in Top 20')"],
    [""],
    ["Tips:"],
    ["  - Focus on one profile at a time (20 programs per profile)"],
    ["  - Use filters to view one profile's rows"],
    ["  - Estimated time: 2-3 hours for all 100 combinations"],
    ["  - Save frequently!"],
    [""],
    ["After labeling:"],
    ["  Run 'test_agent3_with_ground_truth.py' to compare Agent3's output with your labels"],
]

for r_idx, row in enumerate(instructions, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws_instructions.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        elif "Sheet" in str(value):
            cell.font = Font(bold=True, size=12)
        elif value and str(value).startswith("  "):
            cell.font = Font(italic=True)

ws_instructions.column_dimensions['A'].width = 30
ws_instructions.column_dimensions['B'].width = 80

# Save workbook
output_file = 'agent3_ground_truth_template.xlsx'
wb.save(output_file)

print(f"\n✅ Excel template created: {output_file}")
print(f"\nSummary:")
print(f"  - 5 profiles × 20 programs = 100 combinations")
print(f"  - Sheet 1: Hard Constraints ({len(hard_constraints_data)} rows)")
print(f"  - Sheet 2: Degree Compatibility ({len(degree_data)} rows)")
print(f"  - Sheet 3: Overall Ranking ({len(ranking_data)} rows)")
print(f"  - Sheet 4: Instructions")
print(f"\nNext steps:")
print(f"  1. Open {output_file}")
print(f"  2. Follow instructions in the first sheet")
print(f"  3. Label each combination (estimated 2-3 hours)")
print(f"  4. Save as 'agent3_ground_truth.xlsx'")
print(f"  5. Run 'test_agent3_with_ground_truth.py'")
