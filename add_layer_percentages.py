"""
Add layer-by-layer percentage calculations to agent3_ground_truth_FILLED.xlsx
This creates summary sheets showing pass rates for each layer and profile
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

def calculate_layer_percentages(input_file, output_file):
    """Calculate pass rates for each layer and profile"""
    
    wb = openpyxl.load_workbook(input_file)
    
    # Read Hard Constraints sheet
    ws_hard = wb['Hard Constraints']
    
    # Data structures to store results
    profile_stats = defaultdict(lambda: {
        'total_programs': 0,
        'gpa_pass': 0,
        'tuition_pass': 0,
        'location_pass': 0,
        'semester_pass': 0,
        'language_pass': 0,
        'work_exp_pass': 0,
        'ects_pass': 0,
        'all_hard_pass': 0
    })
    
    # Get column indices
    headers = [cell.value for cell in ws_hard[1]]
    col_profile = headers.index('Profile ID') + 1
    col_gpa = headers.index('GPA Pass?') + 1
    col_tuition = headers.index('Tuition Pass?') + 1
    col_location = headers.index('Location Pass?') + 1
    col_semester = headers.index('Semester Pass?') + 1
    
    # Find additional columns if they exist
    col_language = headers.index('Language Pass?') + 1 if 'Language Pass?' in headers else None
    col_work_exp = headers.index('Work Experience Pass?') + 1 if 'Work Experience Pass?' in headers else None
    col_ects = headers.index('ECTS Pass?') + 1 if 'ECTS Pass?' in headers else None
    
    # Process each row
    for row in ws_hard.iter_rows(min_row=2, max_row=ws_hard.max_row):
        profile_id = row[col_profile - 1].value
        if not profile_id:
            continue
            
        stats = profile_stats[profile_id]
        stats['total_programs'] += 1
        
        # Count passes for each layer
        if row[col_gpa - 1].value == 'YES':
            stats['gpa_pass'] += 1
        if row[col_tuition - 1].value == 'YES':
            stats['tuition_pass'] += 1
        if row[col_location - 1].value == 'YES':
            stats['location_pass'] += 1
        if row[col_semester - 1].value == 'YES':
            stats['semester_pass'] += 1
            
        if col_language and row[col_language - 1].value == 'YES':
            stats['language_pass'] += 1
        if col_work_exp and row[col_work_exp - 1].value == 'YES':
            stats['work_exp_pass'] += 1
        if col_ects and row[col_ects - 1].value == 'YES':
            stats['ects_pass'] += 1
            
        # Check if all hard constraints pass
        all_pass = (
            row[col_gpa - 1].value == 'YES' and
            row[col_tuition - 1].value == 'YES' and
            row[col_location - 1].value == 'YES' and
            row[col_semester - 1].value == 'YES'
        )
        if col_language:
            all_pass = all_pass and row[col_language - 1].value == 'YES'
        if col_work_exp:
            all_pass = all_pass and row[col_work_exp - 1].value == 'YES'
        if col_ects:
            all_pass = all_pass and row[col_ects - 1].value == 'YES'
            
        if all_pass:
            stats['all_hard_pass'] += 1
    
    # Read Degree Compatibility sheet
    ws_degree = wb['Degree Compatibility']
    degree_stats = defaultdict(lambda: {'total': 0, 'compatible': 0})
    
    headers_deg = [cell.value for cell in ws_degree[1]]
    col_profile_deg = headers_deg.index('Profile ID') + 1
    col_compatible = headers_deg.index('Should Match?') + 1
    
    for row in ws_degree.iter_rows(min_row=2, max_row=ws_degree.max_row):
        profile_id = row[col_profile_deg - 1].value
        if not profile_id:
            continue
        degree_stats[profile_id]['total'] += 1
        if row[col_compatible - 1].value == 'YES':
            degree_stats[profile_id]['compatible'] += 1
    
    # Read Overall Ranking sheet
    ws_ranking = wb['Overall Ranking']
    ranking_stats = defaultdict(lambda: {'total': 0, 'expected': 0})
    
    headers_rank = [cell.value for cell in ws_ranking[1]]
    col_profile_rank = headers_rank.index('Profile ID') + 1
    col_expected = headers_rank.index('Expected Top N') + 1
    
    for row in ws_ranking.iter_rows(min_row=2, max_row=ws_ranking.max_row):
        profile_id = row[col_profile_rank - 1].value
        if not profile_id:
            continue
        ranking_stats[profile_id]['total'] += 1
        # Check if Expected Top N has a value (not empty)
        expected_val = row[col_expected - 1].value
        if expected_val and str(expected_val).strip():
            ranking_stats[profile_id]['expected'] += 1
    
    # Create new summary sheet
    if 'Layer Summary' in wb.sheetnames:
        del wb['Layer Summary']
    ws_summary = wb.create_sheet('Layer Summary', 0)
    
    # Style definitions
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    ws_summary['A1'] = 'Layer-by-Layer Pass Rate Analysis'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:M1')
    
    # Column headers
    headers = [
        'Profile ID',
        'Total Programs',
        'Layer 1: Hard Constraints Pass',
        'Layer 1: Hard Constraints %',
        'Layer 2: Degree Compatible',
        'Layer 2: Degree %',
        'Expected in Top 10',
        'Match Rate %'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Write data for each profile
    row_num = 4
    for profile_id in sorted(profile_stats.keys()):
        stats = profile_stats[profile_id]
        total = stats['total_programs']
        
        deg_stats = degree_stats.get(profile_id, {'total': 0, 'compatible': 0})
        rank_stats = ranking_stats.get(profile_id, {'total': 0, 'expected': 0})
        
        data = [
            profile_id,
            total,
            stats['all_hard_pass'],
            f"{stats['all_hard_pass']/total*100:.1f}%" if total > 0 else "N/A",
            deg_stats['compatible'],
            f"{deg_stats['compatible']/deg_stats['total']*100:.1f}%" if deg_stats['total'] > 0 else "N/A",
            rank_stats['expected'],
            f"{rank_stats['expected']/rank_stats['total']*100:.1f}%" if rank_stats['total'] > 0 else "N/A"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_summary.cell(row=row_num, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Highlight percentage cells (columns 4, 6, 8)
            if col in [4, 6, 8]:
                if isinstance(value, str) and '%' in value:
                    pct = float(value.rstrip('%'))
                    if pct >= 80:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif pct >= 50:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        row_num += 1
    
    # Add average row
    ws_summary.cell(row=row_num + 1, column=1, value='AVERAGE').font = Font(bold=True)
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 28
    ws_summary.column_dimensions['D'].width = 28
    ws_summary.column_dimensions['E'].width = 25
    ws_summary.column_dimensions['F'].width = 18
    ws_summary.column_dimensions['G'].width = 20
    ws_summary.column_dimensions['H'].width = 15
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Layer percentages added to {output_file}")
    print(f"✓ Created 'Layer Summary' sheet with pass rates for each layer")
    
    # Print summary
    print("\n" + "="*80)
    print("LAYER-BY-LAYER PASS RATES SUMMARY")
    print("="*80)
    for profile_id in sorted(profile_stats.keys()):
        stats = profile_stats[profile_id]
        total = stats['total_programs']
        print(f"\n{profile_id}:")
        print(f"  Layer 1 (Hard Constraints): {stats['all_hard_pass']:2d}/{total} ({stats['all_hard_pass']/total*100:5.1f}%)")
        
        deg_stats = degree_stats.get(profile_id, {'total': 0, 'compatible': 0})
        if deg_stats['total'] > 0:
            print(f"  Layer 2 (Degree Match):     {deg_stats['compatible']:2d}/{deg_stats['total']} ({deg_stats['compatible']/deg_stats['total']*100:5.1f}%)")
        
        rank_stats = ranking_stats.get(profile_id, {'total': 0, 'expected': 0})
        if rank_stats['total'] > 0:
            print(f"  Match Rate:                 {rank_stats['expected']:2d}/{rank_stats['total']} ({rank_stats['expected']/rank_stats['total']*100:5.1f}%)")

if __name__ == '__main__':
    input_file = 'agent3_ground_truth_FILLED.xlsx'
    output_file = 'agent3_ground_truth_FILLED.xlsx'  # Overwrite same file
    calculate_layer_percentages(input_file, output_file)
